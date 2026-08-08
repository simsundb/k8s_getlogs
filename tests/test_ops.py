# tests/test_ops.py
"""第④页 SSH 运维：预置命令、命令构建、表格解析、HTML/Excel 导出、运维项持久化。"""
from openpyxl import load_workbook

from src.ops import (OPS_COMMANDS, OpsCommand, OpsResult, build_command,
                     export_excel, export_html, load_ops_commands,
                     parse_table_output, save_ops_commands)


# ---------- 预置命令 ----------
def test_ops_commands_well_formed():
    labels = [c.label for c in OPS_COMMANDS]
    assert len(labels) == len(set(labels))            # 名称唯一
    cats = {c.category for c in OPS_COMMANDS}
    assert cats <= {"集群", "节点", "应用", "存储"}
    assert len(OPS_COMMANDS) >= 15                     # 常规运维项数量足够
    for c in OPS_COMMANDS:
        assert c.command, c.label
        if "{namespace}" in c.command:
            assert c.needs_namespace, c.label         # 用命名空间的命令必须标记


def test_build_command_replaces_namespace():
    cmd = build_command("kubectl get pods -n {namespace} -o wide", "ns-1")
    assert cmd == "kubectl get pods -n ns-1 -o wide"


def test_build_command_without_placeholder_unchanged():
    cmd = build_command("kubectl cluster-info", "ns-1")
    assert cmd == "kubectl cluster-info"


def test_build_command_empty_namespace_uses_default():
    cmd = build_command("kubectl get pods -n {namespace}", "")
    assert cmd == "kubectl get pods -n default"


# ---------- 表格解析 ----------
def test_parse_table_output_aligned_columns():
    text = "NAME    STATUS   ROLES\nnode-1  Ready    master\nnode-2  Ready    worker\n"
    rows = parse_table_output(text)
    assert rows == [
        ["NAME", "STATUS", "ROLES"],
        ["node-1", "Ready", "master"],
        ["node-2", "Ready", "worker"],
    ]


def test_parse_table_output_single_word_line_kept_whole():
    text = "hello\nthis is a single spaced line\n"
    rows = parse_table_output(text)
    assert rows == [["hello"], ["this is a single spaced line"]]


def test_parse_table_output_ignores_blank_lines():
    assert parse_table_output("a  b\n\n\nc  d\n") == [["a", "b"], ["c", "d"]]


# ---------- HTML 导出 ----------
def _sample_results():
    return [
        OpsResult(label="节点列表", command="kubectl get nodes",
                  start_time="2026-08-08 10:00:00", ok=True,
                  output="NAME     STATUS\nnode-1    Ready", exit_code=0, duration=1.2),
        OpsResult(label="集群信息", command="kubectl cluster-info",
                  start_time="2026-08-08 10:00:05", ok=False,
                  error="error: no context", exit_code=1, duration=0.8),
    ]


def test_export_html_writes_utf8_file(tmp_path):
    out = tmp_path / "report.html"
    path = export_html(_sample_results(), out)
    assert path == str(out)
    text = out.read_text(encoding="utf-8")
    assert "SSH 运维结果" in text
    assert "节点列表" in text and "kubectl get nodes" in text
    assert "node-1" in text
    assert "失败" in text and "error: no context" in text
    assert 'charset="utf-8"' in text


# ---------- Excel 导出 ----------
def test_export_excel_writes_summary_and_per_result_sheets(tmp_path):
    out = tmp_path / "report.xlsx"
    export_excel(_sample_results(), out)
    assert out.exists()

    wb = load_workbook(out)
    assert wb.sheetnames == ["汇总", "1-节点列表", "2-集群信息"]

    ws = wb["汇总"]
    header = [c.value for c in ws[1]]
    assert header == ["序号", "运维项", "状态", "退出码", "耗时(秒)", "输出行数", "命令"]
    assert [c.value for c in ws[2]] == [1, "节点列表", "成功", 0, 1.2, 2,
                                        "kubectl get nodes"]

    detail = wb["1-节点列表"]
    assert detail["A1"].value == "运维项"
    assert detail["B1"].value == "节点列表"
    # 解析出的表头行落在明细工作表中
    assert detail["A9"].value == "NAME"
    assert detail["B9"].value == "STATUS"
    assert detail["A10"].value == "node-1"


# ---------- 运维项持久化（config.json ops_commands） ----------
def test_load_ops_commands_defaults_when_missing():
    """未保存过（首次运行）→ 返回出厂预置，且全部启用。"""
    cmds = load_ops_commands({})
    assert len(cmds) == len(OPS_COMMANDS)
    assert all(c.active for c in cmds)


def test_load_ops_commands_empty_list_falls_back_to_defaults():
    assert len(load_ops_commands({"ops_commands": []})) == len(OPS_COMMANDS)


def test_save_and_load_roundtrip(monkeypatch, tmp_path):
    """保存的新增/停用项能完整读回。"""
    from src import config
    monkeypatch.setattr(config, "CONFIG_PATH", tmp_path / "config.json")
    monkeypatch.setattr(config, "APP_DIR", tmp_path)

    cmds = [OpsCommand("磁盘占用", "节点", "Master 磁盘", "df -h", False, True),
            OpsCommand("自定义检查", "自定义", "检查服务", "echo ok", False, False)]
    config.save_config(config.load_config() | save_ops_commands({}, cmds))
    loaded = load_ops_commands(config.load_config())
    assert [c.label for c in loaded] == ["磁盘占用", "自定义检查"]
    assert loaded[0].active is True
    assert loaded[1].active is False
    assert loaded[1].category == "自定义"


def test_load_ops_commands_skips_bad_entries(monkeypatch, tmp_path):
    """非法条目（缺 label）被跳过，合法条目保留。"""
    from src import config
    monkeypatch.setattr(config, "CONFIG_PATH", tmp_path / "config.json")
    monkeypatch.setattr(config, "APP_DIR", tmp_path)
    config.save_config(config.load_config() | {"ops_commands": [
        {"label": "", "command": "echo 1"},                          # 非法：无名称
        {"label": "合法项", "category": "存储", "description": "d",
         "command": "kubectl get pv", "needs_namespace": False, "active": True},
    ]})
    loaded = load_ops_commands(config.load_config())
    assert [c.label for c in loaded] == ["合法项"]
