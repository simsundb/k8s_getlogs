"""SSH 运维中心：预置运维命令、命令构建、结果导出（HTML / Excel）。

预置命令等价于「手工 SSH 到 MASTER 后敲一条命令」的常见日常运维操作；
命令模板中可含 {namespace} 占位符，执行时替换为当前选中命名空间。
"""
import html
import re
from dataclasses import dataclass

# 列分隔：2 个及以上空格 / 制表符（kubectl、df 等对齐输出以多空格分列）
_TABLE_SEP = re.compile(r"[ \t]{2,}")


@dataclass
class OpsCommand:
    label: str                 # 运维项名称
    category: str              # 类别：集群 / 节点 / 应用 / 存储
    description: str           # 中文说明
    command: str               # 实际命令模板（可含 {namespace}）
    needs_namespace: bool = False  # 是否依赖当前选中命名空间


OPS_COMMANDS: list[OpsCommand] = [
    # ---------- 集群 ----------
    OpsCommand("集群信息", "集群", "查看 API 服务地址、版本等集群基本信息",
               "kubectl cluster-info"),
    OpsCommand("集群组件状态", "集群", "查看控制面组件 scheduler / controller-manager 健康状态",
               "kubectl get cs"),
    OpsCommand("命名空间列表", "集群", "列出全部命名空间",
               "kubectl get namespaces"),
    OpsCommand("节点列表", "集群", "查看所有节点地址、状态、K8s 版本、OS",
               "kubectl get nodes -o wide"),
    OpsCommand("节点资源占用", "集群", "查看节点 CPU / 内存实时使用率（需 metrics-server）",
               "kubectl top nodes"),
    # ---------- 节点 ----------
    OpsCommand("节点磁盘占用", "节点", "Master 节点磁盘分区使用情况",
               "df -h"),
    OpsCommand("节点内存使用", "节点", "Master 节点内存使用情况",
               "free -h"),
    OpsCommand("节点负载与运行时长", "节点", "1 / 5 / 15 分钟平均负载与开机时长",
               "uptime"),
    OpsCommand("节点系统信息", "节点", "内核、主机名、硬件架构",
               "uname -a"),
    OpsCommand("内存占用 TOP 进程", "节点", "按内存占用排序的前 20 个进程",
               "ps aux --sort=-%mem | head -20"),
    # ---------- 应用 ----------
    OpsCommand("命名空间下全部 Pod", "应用", "查看所选命名空间全部 Pod 状态与节点分布",
               "kubectl get pods -n {namespace} -o wide", needs_namespace=True),
    OpsCommand("全部命名空间 Pod", "应用", "跨命名空间查看所有 Pod 状态",
               "kubectl get pods -A -o wide"),
    OpsCommand("命名空间下 Deployment", "应用", "查看部署期望副本 / 可用副本数",
               "kubectl get deployments -n {namespace}", needs_namespace=True),
    OpsCommand("Service / Endpoints", "应用", "查看所选命名空间服务与端点",
               "kubectl get svc,endpoints -n {namespace}", needs_namespace=True),
    OpsCommand("Pod 资源占用 TOP", "应用", "所选命名空间 Pod CPU / 内存使用（需 metrics-server）",
               "kubectl top pods -n {namespace}", needs_namespace=True),
    OpsCommand("最近事件", "应用", "所选命名空间最近 50 条事件（按发生时间排序）",
               "kubectl get events -n {namespace} --sort-by=.lastTimestamp | tail -50",
               needs_namespace=True),
    OpsCommand("ConfigMap / Secret", "应用", "所选命名空间配置项与密钥（仅列名称，不打印值）",
               "kubectl get cm,secrets -n {namespace}", needs_namespace=True),
    # ---------- 存储 ----------
    OpsCommand("PV / PVC 列表", "存储", "查看持久卷与持久卷声明及其状态",
               "kubectl get pv,pvc -A"),
]


def build_command(command: str, namespace: str = "") -> str:
    """把命令模板中的 {namespace} 替换为命名空间；无占位符原样返回。"""
    if "{namespace}" not in command:
        return command
    return command.format(namespace=namespace or "default")


@dataclass
class OpsResult:
    label: str
    command: str
    start_time: str            # 执行时刻，如 2026-08-08 12:00:00
    ok: bool
    output: str = ""
    error: str = ""
    exit_code: int = -1
    duration: float = 0.0      # 秒

    @property
    def status(self) -> str:
        return "成功" if self.ok else "失败"

    def display(self) -> str:
        """输出面板中的完整回显文本。"""
        head = (f"===== [{self.label}] {self.start_time} ｜ 退出码 {self.exit_code}"
                f" ｜ 耗时 {self.duration:.1f}s ｜ {self.status} =====")
        body = (self.output.strip() if self.ok and self.output.strip()
                else self.error.strip())
        if not body:
            body = "(无输出)"
        return "\n".join([head, f"$ {self.command}", body])


def parse_table_output(text: str) -> list[list[str]]:
    """把对齐文本拆成表格行：按 2+ 空格/制表符分列，单列行保持整行。

    kubectl get / df / free 等输出列间以多空格对齐 → 拆成多列；
    ps 等单空格分隔的行 → 整行单列，保证内容不丢。
    """
    rows: list[list[str]] = []
    for line in text.splitlines():
        if not line.strip():
            continue
        cells = [c for c in _TABLE_SEP.split(line.strip()) if c]
        rows.append(cells if len(cells) >= 2 else [line.strip()])
    return rows


# ---------------- 导出 ----------------

def _result_section(res: OpsResult, idx: int) -> str:
    """单个结果的 HTML 区块。"""
    status_color = "#27ae60" if res.ok else "#c0392b"
    body = html.escape(res.output) if res.ok else html.escape(res.error or res.output)
    body = body or "(无输出)"
    return f"""
    <div class="card">
      <div class="card-head">
        <span class="badge" style="background:{status_color}">{res.status}</span>
        <span class="label">{html.escape(res.label)}</span>
        <span class="meta">#{idx} ｜ {html.escape(res.start_time)} ｜ 退出码 {res.exit_code} ｜ 耗时 {res.duration:.1f}s</span>
      </div>
      <pre class="cmd">$ {html.escape(res.command)}</pre>
      <pre class="out">{body}</pre>
    </div>"""


def export_html(results: list[OpsResult], path) -> str:
    """把结果列表导出为独立 HTML 文件，返回写出的文件路径。"""
    ok = sum(1 for r in results if r.ok)
    failed = len(results) - ok
    cards = "\n".join(_result_section(r, i + 1) for i, r in enumerate(results))
    doc = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>SSH 运维结果</title>
<style>
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:#f3f5f9; margin:20px; color:#2b3240; }}
  h1 {{ font-size:20px; }}
  .summary {{ color:#6a7380; margin-bottom:16px; }}
  .card {{ background:#fff; border:1px solid #d8dce4; border-radius:8px; padding:12px 16px; margin-bottom:14px; }}
  .card-head {{ display:flex; align-items:center; gap:10px; margin-bottom:8px; }}
  .badge {{ color:#fff; border-radius:4px; padding:1px 8px; font-size:12px; }}
  .label {{ font-weight:bold; font-size:15px; }}
  .meta {{ color:#6a7380; font-size:12px; }}
  pre {{ margin:4px 0; padding:8px 10px; border-radius:4px; font-size:12px; overflow-x:auto; white-space:pre; }}
  pre.cmd {{ background:#edf0f6; color:#2f5ca8; }}
  pre.out {{ background:#f8fafd; border:1px solid #e9ecf2; max-height:520px; overflow-y:auto; }}
</style>
</head>
<body>
  <h1>SSH 运维结果</h1>
  <div class="summary">共 {len(results)} 条：成功 {ok}，失败 {failed}，生成于 {results[0].start_time if results else "-"}</div>
  {cards}
</body>
</html>
"""
    path = str(path)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(doc)
    return path


def _sheet_name(idx: int, label: str) -> str:
    """Excel 工作表名：去掉非法字符，最长 31 字符，带序号避免重复。"""
    cleaned = re.sub(r"[\[\]:*?/\\]", "", label)
    name = f"{idx}-{cleaned}"
    return name[:31] or f"result-{idx}"


def export_excel(results: list[OpsResult], path) -> str:
    """把结果列表导出为 .xlsx：一个「汇总」表 + 每个结果一个工作表。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(["序号", "运维项", "状态", "退出码", "耗时(秒)", "输出行数", "命令"])
    head_fill = PatternFill("solid", fgColor="3B6FC4")
    head_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
    for i, r in enumerate(results, 1):
        out_lines = len(r.output.splitlines()) if r.output else 0
        ws.append([i, r.label, r.status, r.exit_code,
                   round(r.duration, 2), out_lines, r.command])
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 70
    ws.freeze_panes = "A2"

    for idx, res in enumerate(results, 1):
        s = wb.create_sheet(_sheet_name(idx, res.label))
        s.append(["运维项", res.label])
        s.append(["命令", res.command])
        s.append(["执行时间", res.start_time])
        s.append(["退出码", res.exit_code])
        s.append(["耗时(秒)", round(res.duration, 2)])
        s.append(["状态", res.status])
        s.append(["输出", ""])
        s.append(["--------", "--------"])
        for col in ("A", "B"):
            s.column_dimensions[col].width = 12 if col == "A" else 100
        for row in s.iter_rows(min_row=1, max_row=1):
            for c in row:
                c.font = Font(bold=True)
        for row in parse_table_output(res.output):
            s.append(row)
        if not res.ok and res.error:
            s.append([res.error])

    path = str(path)
    wb.save(path)
    return path
